import os
import datetime
from sqlalchemy.orm import Session
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from backend import models

def generate_stock_pdf(db: Session, filepath: str) -> str:
    """Skladdagi joriy qoldiqlar ro'yxatini PDF shaklida yaratadi"""
    # Create directory if not exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    doc = SimpleDocTemplate(
        filepath, 
        pagesize=letter, 
        rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Styles for Premium Look
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#1E293B'), # Slate 800
        spaceAfter=6,
        alignment=1 # Center
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#64748B'), # Slate 500
        spaceAfter=15,
        alignment=1
    )
    
    table_text_style = ParagraphStyle(
        'TableText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#334155')
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )
    
    elements = []
    
    # Title & Subtitle
    elements.append(Paragraph("BARAKA SKLAD — ERP TIZIMI", title_style))
    now_str = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    elements.append(Paragraph(f"Joriy Ombor Qoldiqlari Hisoboti — Chop etilgan vaqt: {now_str}", subtitle_style))
    elements.append(Spacer(1, 10))
    
    # Fetch Data from DB
    stocks = db.query(models.Stock).join(models.Product).all()
    
    # Table data header
    data = [
        [
            Paragraph("T/r", table_header_style),
            Paragraph("Shtrix-kod", table_header_style),
            Paragraph("Tovar Nomi", table_header_style),
            Paragraph("Toifasi", table_header_style),
            Paragraph("Ombor", table_header_style),
            Paragraph("Qoldiq", table_header_style),
            Paragraph("Sotish Narxi", table_header_style),
            Paragraph("Status", table_header_style)
        ]
    ]
    
    for idx, s in enumerate(stocks, 1):
        # Determine status
        status = "Normal"
        status_color = "#10B981" # Green
        if s.quantity <= s.product.min_threshold:
            status = "Kam qoldi!"
            status_color = "#EF4444" # Red
        
        status_style = ParagraphStyle(
            f'Status_{idx}',
            parent=table_text_style,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor(status_color)
        )
        
        data.append([
            Paragraph(str(idx), table_text_style),
            Paragraph(s.product.barcode, table_text_style),
            Paragraph(s.product.name, table_text_style),
            Paragraph(s.product.category.name, table_text_style),
            Paragraph(s.warehouse.name, table_text_style),
            Paragraph(f"{s.quantity}", table_text_style),
            Paragraph(f"{s.product.selling_price:,.0f} UZS", table_text_style),
            Paragraph(status, status_style)
        ])
    
    # Create Table
    # Widths sum to 552 (letter width is 612 - 60 margin = 552)
    col_widths = [25, 80, 140, 90, 80, 45, 52, 40]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Table Styling
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F172A')), # Dark Slate header
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
    ])
    
    # Alternating row background
    for i in range(1, len(data)):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0,i), (-1,i), colors.HexColor('#F8FAFC'))
            
    t.setStyle(t_style)
    elements.append(t)
    
    doc.build(elements)
    return filepath

def generate_accounting_excel(db: Session, filepath: str) -> str:
    """Barcha moliyaviy va kirim-chiqim ma'lumotlarini Excel formatida yaratadi (Multi-sheet)"""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb = Workbook()
    
    # Color palette (Navy / Steel Blue)
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    title_font = Font(name='Arial', size=14, bold=True, color='1E293B')
    regular_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    accent_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    warning_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # --- SHEET 1: OMBOB QOLDIQLARI ---
    ws1 = wb.active
    ws1.title = "Ombor Qoldiqlari"
    
    # Title
    ws1.append(["BARAKA SKLAD ERP — JORIY QOLDIQLAR"])
    ws1.cell(1, 1).font = title_font
    ws1.append([f"Chop etilgan sana: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"])
    ws1.append([]) # empty row
    
    # Table headers
    headers1 = ["T/r", "Shtrix-kod", "Tovar Nomi", "Kategoriya", "Ombor", "Qoldiq", "Tannarxi (UZS)", "Sotish Narxi (UZS)", "Jami Tannarxi", "Jami Sotish Narxi", "Chegara", "Status"]
    ws1.append(headers1)
    
    for col_idx, header in enumerate(headers1, 1):
        cell = ws1.cell(4, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    stocks = db.query(models.Stock).join(models.Product).all()
    
    for idx, s in enumerate(stocks, 1):
        status = "Normal"
        is_low = s.quantity <= s.product.min_threshold
        if is_low:
            status = "Kam qoldi!"
            
        row_data = [
            idx,
            s.product.barcode,
            s.product.name,
            s.product.category.name,
            s.warehouse.name,
            s.quantity,
            s.product.cost_price,
            s.product.selling_price,
            s.quantity * s.product.cost_price,
            s.quantity * s.product.selling_price,
            s.product.min_threshold,
            status
        ]
        ws1.append(row_data)
        
        row_num = ws1.max_row
        # Styling cells
        for col_idx in range(1, len(row_data) + 1):
            cell = ws1.cell(row_num, col_idx)
            cell.font = regular_font
            cell.border = thin_border
            
            # Alignments & Formats
            if col_idx in [1, 2, 5, 12]:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx in [6, 11]:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00'
            elif col_idx in [7, 8, 9, 10]:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
                
            # Zebra pattern & warning
            if is_low and col_idx == 12:
                cell.fill = warning_fill
                cell.font = Font(name='Arial', size=10, bold=True, color='991B1B')
            elif row_num % 2 == 0:
                cell.fill = zebra_fill

    # --- SHEET 2: KINIM-CHIQIM TRANSAKSIYALARI ---
    ws2 = wb.create_sheet("Tranzaksiyalar Tarixi")
    ws2.append(["AMALLAR TAHRIRI (TRANSAKSIYALAR)"])
    ws2.cell(1, 1).font = title_font
    ws2.append([f"Davr: Barcha davrlar ({datetime.datetime.now().strftime('%d.%m.%Y')})"])
    ws2.append([])
    
    headers2 = ["ID", "Vaqt", "Tovar Nomi", "Amal Turi", "Sklad", "Miqdor", "Narxi (UZS)", "Jami Summa (UZS)", "Mijoz (Usta)", "Operator (Skladchi)"]
    ws2.append(headers2)
    
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(4, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    txs = db.query(models.Transaction).order_by(models.Transaction.created_at.desc()).all()
    
    for idx, t in enumerate(txs, 1):
        client_name = t.customer.name if t.customer else "—"
        user_name = t.operator.full_name if t.operator else "Tizim"
        
        # Calculate price based on transaction type
        price = t.selling_price if t.type in ["Chiqim", "Transfer"] else t.cost_price
        total_sum = t.quantity * price
        
        row_data = [
            t.id,
            t.created_at.strftime("%d.%m.%Y %H:%M"),
            t.product.name,
            t.type,
            t.warehouse.name,
            t.quantity,
            price,
            total_sum,
            client_name,
            user_name
        ]
        ws2.append(row_data)
        
        row_num = ws2.max_row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws2.cell(row_num, col_idx)
            cell.font = regular_font
            cell.border = thin_border
            
            if col_idx in [1, 2, 4, 5]:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00'
            elif col_idx in [7, 8]:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
                
            if row_num % 2 == 0:
                cell.fill = zebra_fill
                
            # Color code types
            if col_idx == 4:
                if t.type == "Kirim":
                    cell.font = Font(name='Arial', size=10, bold=True, color='047857') # Green
                elif t.type == "Chiqim":
                    cell.font = Font(name='Arial', size=10, bold=True, color='B91C1C') # Red
                elif t.type == "Transfer":
                    cell.font = Font(name='Arial', size=10, bold=True, color='1D4ED8') # Blue

    # --- SHEET 3: USTALAR (MIJOZLAR) ---
    ws3 = wb.create_sheet("Ustalar Ro'yxati")
    ws3.append(["USTALAR VA MIJOZLAR BALANCE VA STATUS"])
    ws3.cell(1, 1).font = title_font
    ws3.append([])
    
    headers3 = ["ID", "Usta F.I.Sh", "Telefon Raqami", "Balans (UZS)", "Holati"]
    ws3.append(headers3)
    
    for col_idx, header in enumerate(headers3, 1):
        cell = ws3.cell(3, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    customers = db.query(models.Customer).all()
    
    for idx, c in enumerate(customers, 1):
        status = "Qarzi yo'q"
        if c.balance < 0:
            status = f"Qarzdor ({abs(c.balance):,.0f} UZS)"
        elif c.balance > 0:
            status = "Oldindan to'lov"
            
        row_data = [
            c.id,
            c.name,
            c.phone or "—",
            c.balance,
            status
        ]
        ws3.append(row_data)
        
        row_num = ws3.max_row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws3.cell(row_num, col_idx)
            cell.font = regular_font
            cell.border = thin_border
            
            if col_idx in [1, 3, 5]:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
                
            if c.balance < 0 and col_idx == 5:
                cell.fill = warning_fill
                cell.font = Font(name='Arial', size=10, bold=True, color='991B1B')
            elif c.balance > 0 and col_idx == 5:
                cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                cell.font = Font(name='Arial', size=10, bold=True, color='065F46')
            elif row_num % 2 == 0:
                cell.fill = zebra_fill

    # Auto-adjust column widths for all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    # check for multi-line or headers
                    val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            col_letter = col[0].column_letter
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
    wb.save(filepath)
    return filepath
